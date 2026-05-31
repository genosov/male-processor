from openpyxl import load_workbook

from extensions.analytics_report import report_generator
from src.mail_processor.models import ProcessingStats


def test_generate_excel_report_uses_same_snapshot_for_categories_and_files(tmp_path, monkeypatch):
    processed_dir = tmp_path / "processed"
    reports_dir = tmp_path / "reports"

    critical_dir = processed_dir / "critical"
    support_dir = processed_dir / "support"
    critical_dir.mkdir(parents=True)
    support_dir.mkdir(parents=True)

    (critical_dir / "mail_0001.txt").write_text("critical", encoding="utf-8")
    (support_dir / "mail_0002.txt").write_text("support", encoding="utf-8")
    (support_dir / "mail_0003.txt").write_text("support", encoding="utf-8")

    monkeypatch.setattr(report_generator, "PROCESSED_DIR", processed_dir)
    monkeypatch.setattr(report_generator, "REPORTS_DIR", reports_dir)
    monkeypatch.setattr(
        report_generator,
        "DEFAULT_REPORT_PATH",
        reports_dir / "processing_report.xlsx",
    )

    output_path = report_generator.generate_excel_report(ProcessingStats())

    workbook = load_workbook(output_path)
    summary_sheet = workbook["Summary"]
    categories_sheet = workbook["Categories"]
    files_sheet = workbook["ProcessedFiles"]

    assert summary_sheet["A9"].value == "Файлов в архиве"
    assert summary_sheet["B9"].value == 3

    assert categories_sheet["B4"].value == 1
    assert categories_sheet["B5"].value == 2
    assert files_sheet.max_row == 6
