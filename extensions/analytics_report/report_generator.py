from datetime import datetime
from pathlib import Path

from src.mail_processor.config import PROCESSED_DIR, ROOT_DIR
from src.mail_processor.file_handler import FileHandler
from src.mail_processor.models import CATEGORIES


REPORTS_DIR = ROOT_DIR / "reports"
DEFAULT_REPORT_PATH = REPORTS_DIR / "processing_report.xlsx"
DATE_FORMAT = "%Y-%m-%d %H:%M:%S"


def generate_excel_report(stats, output_path=None) -> Path:
    """Create an Excel report with processing statistics and file metadata."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = Path(output_path) if output_path else DEFAULT_REPORT_PATH
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    processed_snapshot = _collect_processed_files_snapshot()

    styles = {
        "title_font": Font(size=16, bold=True),
        "header_font": Font(bold=True),
        "header_fill": PatternFill("solid", fgColor="D9EAF7"),
        "border": Border(
            left=Side(style="thin", color="808080"),
            right=Side(style="thin", color="808080"),
            top=Side(style="thin", color="808080"),
            bottom=Side(style="thin", color="808080"),
        ),
        "center": Alignment(horizontal="center"),
    }

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _fill_summary_sheet(summary_sheet, stats, processed_snapshot)

    categories_sheet = workbook.create_sheet("Categories")
    _fill_categories_sheet(categories_sheet, processed_snapshot)

    files_sheet = workbook.create_sheet("ProcessedFiles")
    _fill_processed_files_sheet(files_sheet, processed_snapshot)

    for sheet in workbook.worksheets:
        _apply_table_style(sheet, styles)
        _auto_fit_columns(sheet)

    workbook.save(output_path)
    return output_path


def _fill_summary_sheet(sheet, stats, processed_snapshot) -> None:
    sheet["A1"] = "Сводный отчёт обработки корпоративной почты"
    sheet.append([])
    sheet.append(["Метрика", "Значение"])

    total_files = stats.total_files
    success_percent = 0
    if total_files:
        success_percent = round(stats.processed_files / total_files * 100, 2)

    rows = [
        ("Дата отчёта", datetime.now().strftime(DATE_FORMAT)),
        ("Всего файлов", stats.total_files),
        ("Обработано успешно", stats.processed_files),
        ("Ошибок", stats.failed_files),
        ("Процент успешности", f"{success_percent}%"),
        ("Файлов в архиве", processed_snapshot["total_files"]),
    ]

    for row in rows:
        sheet.append(row)


def _fill_categories_sheet(sheet, processed_snapshot) -> None:
    sheet["A1"] = "Распределение писем по категориям"
    sheet.append([])
    sheet.append(["Категория", "Количество"])

    for category in CATEGORIES:
        count = processed_snapshot["categories_count"].get(category, 0)
        sheet.append([category, count])


def _fill_processed_files_sheet(sheet, processed_snapshot) -> None:
    sheet["A1"] = "Обработанные файлы"
    sheet.append([])
    sheet.append(
        [
            "Название файла",
            "Категория",
            "Размер, байт",
            "Расширение",
            "Создан",
            "Изменён",
            "Полный путь",
        ]
    )

    for row in processed_snapshot["rows"]:
        sheet.append(row)


def _collect_processed_files_snapshot() -> dict:
    file_handler = FileHandler()
    rows = []
    categories_count = {category: 0 for category in CATEGORIES}

    for category in CATEGORIES:
        category_dir = PROCESSED_DIR / category
        if not category_dir.exists():
            continue

        for file_path in sorted(category_dir.iterdir()):
            if not file_path.is_file():
                continue

            metadata = file_handler.get_file_metadata(file_path)
            categories_count[category] += 1
            rows.append(
                [
                    metadata["filename"],
                    category,
                    metadata["size_bytes"],
                    metadata["extension"],
                    _format_timestamp(metadata["created_at"]),
                    _format_timestamp(metadata["modified"]),
                    metadata["path"],
                ]
            )

    return {
        "rows": rows,
        "categories_count": categories_count,
        "total_files": len(rows),
    }


def _format_timestamp(timestamp) -> str:
    return datetime.fromtimestamp(timestamp).strftime(DATE_FORMAT)


def _apply_table_style(sheet, styles) -> None:
    sheet["A1"].font = styles["title_font"]
    sheet.freeze_panes = "A4"

    header_row = 3
    for cell in sheet[header_row]:
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.border = styles["border"]
        cell.alignment = styles["center"]

    for row in sheet.iter_rows(min_row=4):
        for cell in row:
            cell.border = styles["border"]


def _auto_fit_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 80)
